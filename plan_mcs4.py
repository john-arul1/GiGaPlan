# -*- coding: utf-8 -*-
"""
Created on Fri Oct 14 16:40:08 2022

@author: John Arul

A program to estimate project time schedule uncertainty

Reads data from an excel file using openpyxl

version - 2

parallel sub activies translated to tree structure

based on defining only successor to tasks

.x is used to define sub tasks

examine edge (n.x,n) to identify subtask brach


node operation = value + max(inputs) + sum(subtask inputs)

TBD:  task graph with time 


version - 4

input sample tagging by adding adictionary to all tasks


"""

import numpy as np
#import pandas as pd
import matplotlib.pyplot as plt
import queue
import openpyxl




##--------------------------------------give input file name here


try:
    
    with open("plan_input_format.txt") as desc:
        data=desc.readlines()

except IOError:
    print("input description file error")
    

if len(data) == 0:
     print ("file is empty")
if len(data) < 6:
     print ("data not correct or less in file")
     
     

dd={}
for line in data[0:6]:
   
   txt,txtval=line.split(":")
   txt=txt.strip()
   txtval=txtval.strip()
   
   if txt.strip() == 'file':
       dd.update({txt:txtval})

   else:
       txtval=int(txtval)
       dd.update({txt:txtval})

nrows=dd['nrows']
ncols=dd['ncols']
srow=dd['srow']
scol=dd['scol']
file=dd['file']
NS=dd['nsamples']   


if NS < 1:
    NS=1    #number of simulations

if NS > 1E8:
   NS=1E8 
   print("WARNING: Too large number of simulations")
   
##---------------------------------------


class cnode:
    def __init__(self,res=0,t=0,suc=0):
        self.res =0
        self.t = 0
        self.suc= 0
        self.prelist=[]
    
    

def sample(p,T):
    
    #cd=reduce(lambda a,b:a+b,p)
    cd=list([ sum(p[:i]) for i in range(len(p))])
    r=np.random.rand()
    
    if r < cd[1]:
        return 0,T[0]
    if (r >= cd[1]) and (r < cd[2]):
        return 1,T[1]
    else: 
        return 2,T[2]
    
    
def get_leafs(nodes,edges):
    leaf=[]
    for i in range(len(nodes)):
        leaf_flag=0
        for j in range(len(edges)):
            if edges[j][1]==nodes[i]:  # if successor of some node
               leaf_flag =1
               
        if leaf_flag==0:
           leaf.append(nodes[i]) 
           
    return leaf       
 
    
# key function 
    
def compute_edge_time(node_time): # uses nodes,edges,    
    #assume DAG
    # BFS bottom to top 
    #make a queue and stack of nodes
    #start from top = -2
    for e in edges:
        if e[1]=='-2':
           head_node = e[0]
           
    qu=queue.Queue()
    st=[]
    qu.put(head_node)
    
    i=0
    while(qu.empty()==False):
        nd=qu.get()
        
        st.append(nd)
        for e in edges:
            if e[1]==nd:
               qu.put(e[0])
    
        i=i+1
        if i > 1E6:           #to be temoved
            break
        
    #print(st) 
    #Extract from stack and compute   
    while(len(st) > 0):
       cn=st.pop()
       
       if cn in leafs:
          #set outgoing edge time
          for e in edges:  # find out edge and update its time as node time
              if cn == e[0]:
                  t=node_time[cn]
                  edge_time.update({e:t})
                  #break;  
    
       else: # if not leaf
           #make list of incoming edges time
           # add one subtask time
           tl=[]
           el=[]
           stl=[]
           for e in edges:
              if cn == e[0]:
                 out_edge=e
              if cn == e[1]:
                 el.append(e)    
           t=0
           #stt=0.0
           for e in el: # incoming edges
              if not sub_task_branch(e):   # returns true for sub task 
                 tl.append(edge_time[e])
              else:
                 stl.append(edge_time[e])
                 #stt= edge_time[e]  # subtask time
                  
           tt=max(tl+[0])+node_time[cn]+max(stl+[0])
           edge_time[out_edge]=tt
           
    
    return edge_time[(head_node,'-2')]


import re


def reverse(s):
    if len(s) == 0:
        return s
    else:
        return reverse(s[1:]) + s[0]


def sub_task_branch(edge):
    
    #return true if it is subtask edge of form (n.x,n)
    e1=edge[0]  # check xl input format numeric or text and simplify
    e2=edge[1]
    
    e1_point=0
    #e2_point=0
    if '.' in e1:
        e1_point =1
        
    e1=reverse(e1)
    
    e1x=re.sub('\d+\.','',e1,1)
    e1x=reverse(e1x)
    
    if(e1x==e2) and e1_point == 1:
        return True
   
    return False 
    

def normalize(Z):
    ss=sum(sum(Z))
    if ss > 1.0E-15:
        Z=Z/ss*Z.shape[1]

    return Z
        

##=======================================    
##======main program


NP =3 # number of probability intervals

    
p=np.ndarray(NP)
T=np.ndarray(NP)



leafs=[]
edges=[]
#edge time is resultant time of node plus input edges
edge_time={} # dictionary of (edge,result)
node_time={} # its sampled time 
node_prob_list={} # possible range
node_time_list={} # possible probabilities


# read excel data into pandas
print ("Processing file : ",file)
#tasks=np.ndarray([5,6]) # 5 = len of nodes
try: 
    wb = openpyxl.load_workbook(file)#,read_only=True) # data_only=True
except IOError:
    print("Excel file open error")

#ws = wb.active
ws = wb.worksheets[0]

#including top header, number of data cells and starting cell 


NR=nrows
NC=ncols
SR=srow
SC=scol


### read data from file

#for row in ws.values:
#    for value in row:
#      print(value)

# for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
# ...   print(row)


# assign data
nodes=[ws.cell(row=SR+1+i, column=SC).value for i in range(NR-1)]

for i in range(len(nodes)): # x.x.x is treated as string  
       nodes[i]=str(nodes[i])


suc_nodes=[ws.cell(row=SR+1+i, column=SC+3).value for i in range(NR-1)]
for i in range(len(suc_nodes)): # x.x.x is treated as string  
       suc_nodes[i]=str(suc_nodes[i])
       if suc_nodes[i].capitalize() =='END':
          suc_nodes[i]=='-2'


wb.close()
##------------------------------
# print data

print("Number of tasks defined: ",len(nodes))
print("List of tasks: ",nodes)
print("List of successor tasks: ",suc_nodes)

##------------------------------

for i in range(len(nodes)):
     
     edges.append((nodes[i],suc_nodes[i]))
     T=[ws.cell(row=SR+1+i, column=SC+4+j).value for j in range(3)]
     T=[float(x) for x in T] 
     p=[ws.cell(row=SR+1+i, column=SC+7+j).value for j in range(3)]
     p=[float(x) for x in p] 
     node_time_list.update({nodes[i]:T}) 
     node_prob_list.update({nodes[i]:p})
     # print(i,T)
     # print(i,p)
leafs=get_leafs(nodes,edges)

print("List of basic tasks: ",leafs)



# start MCS

#ft=np.ndarray(NS)
ft=np.zeros(NS)

#dictionary for back tag
back_tag={}# (node,time,final time slot):count

for i in range(NS):  # number of sample repeatitions    
     sampl_ind=np.zeros(len(nodes))
     for j in range(len(nodes)):  # number of tasks  sample node values
         p=node_prob_list[nodes[j]]
         T=node_time_list[nodes[j]]  # extract T vector
         sampl_ind[j],sv=sample(p,T)
         node_time.update({nodes[j]:sv})  #each node is given a time
        
     ft[i]=compute_edge_time(node_time) # uses nodes,edges,   compute edge values  
     
     # back tagging
     for j in range(len(nodes)):
         if i==0:
             back_tag.update({(nodes[j],sampl_ind[j],int(ft[i])):1})
         
         else: 
            if (nodes[j],sampl_ind[j],int(ft[i])) in back_tag:
                val= back_tag[(nodes[j],sampl_ind[j],int(ft[i]))]
                back_tag.update({(nodes[j],sampl_ind[j],int(ft[i])):val+1})
            else:
               back_tag.update({(nodes[j],sampl_ind[j],int(ft[i])):1})

                  
## ------------------plot and save probability data

Pr=ft/sum(ft)    
    
y,x=np.histogram(ft,10)

x=x[:-1]
dx=np.diff(x)[0]
x=x+dx/2
y=y/sum(y)    

cy=[sum(y[:i+1]) for i in range(len(y)-1)]
cy.append(1.0)

print(x,cy)


fig = plt.figure()

ax1 = fig.add_subplot(121) #
ax2 = fig.add_subplot(122)

#ax1 = fig.add_subplot(121,projection='3d')
ax1.bar(x,y,width=dx)
ax1.set_xlabel("Time")
ax1.set_ylabel("Probability")    
    
plt.savefig('histo_pdf.png')


# plt.bar(x,cy,width=dx)
# plt.xlabel("Time")
# plt.ylabel("Cumulative Probability")    
    
# plt.savefig('histo_cdf.png')

###---------------------------------------------------------------
#  use excel sheet 2 to write prob and plot  

from openpyxl.chart import (BarChart, ScatterChart, 
 Reference, Series, 
 BarChart3D,
 SurfaceChart3D
 )
 
 
chart1 = BarChart()
chart2 = ScatterChart()


#open existing file
#wb=openpyxl.load_workbook(file)
#create a workbook
#wb=openpyxl.Workbook()


if len(wb.worksheets) > 1:
     for sheet in wb.worksheets[1:]:
         wb.remove(sheet)

wb.create_sheet()
ws1=wb.worksheets[1]
ws1.title="Results"

# clear images
# for i in range(len(ws1._images)):
#     print("deleting image",i+1)
#     del ws1._images[i]
    
# clear images
for i in range(len(ws1._charts)):
    print("deleting image",i+1)
    del ws1._charts[i]    
    
#wb.save(file)
# rows = [
#     ('Number', 'Batch 1', 'Batch 2'),
#     (2, 10, 30),
#     (3, 40, 60),
#     (4, 50, 70),
#     (5, 20, 10),
#     (6, 10, 40),
#     (7, 50, 30),
# ]


# for row in rows:
#     ws.append(row)

ws1['A1']="Time"
ws1['B1']="Probability"
ws1['C1']="Cumulative Probability"
for row in range(len(x)):
    ws1[ 'A{}' . format ( row +2)] . value = x[row]
    ws1[ 'B{}' . format ( row +2)] . value = y[row]
    ws1[ 'C{}' . format ( row +2)] . value = cy[row]

        
chart1 . title = "Completion Time Distribution"
chart1 . y_axis . title = 'Probability'
chart1 . x_axis . title = 'Time'
# Now we will create a reference to the data and append the data to the chart.
#wb.save(file)

xval = Reference (ws1 , min_row = 2 , max_row = 1+len(x) , min_col = 1 , max_col = 1 )
yval= Reference (ws1 , min_row = 2 , max_row = 1+len(x) , min_col = 2 , max_col = 2 )
ycval= Reference (ws1 , min_row = 2 , max_row = 1+len(x) , min_col = 3 , max_col = 3 )

s2 = Series(ycval, xvalues=xval)

chart1.type = "col"
chart1.style = 1
chart1.legend=None

chart1.add_data(yval,from_rows=False,titles_from_data=True)
chart1.set_categories(xval)

chart2.append(s2)

chart2.legend=None

# Finally, Add the chart to the sheet and save the file.
ws1. add_chart ( chart1 , "G6" )
ws1. add_chart ( chart2 , "P6" )


####----------------



#analyze back_tag
# create a new sheet
wb.create_sheet()
ws2=wb.worksheets[2]
ws2.title="Inference Plots"


ws2=wb.worksheets[2]


# # clear images
# for i in range(len(ws2._c)):
#     print("deleting image",i+1)
#     del ws2._images[i]
 
#for img in ws._images:   # not working
#     del img    

##-------------------------------------------node slice
node_slice=0  # 0 -> event slice

if node_slice==1:    
    node_select='3'
    tlist=node_time_list[node_select]
    
    #make a set fot ft
    set_ft=set(ft)
    set_ft=list(set_ft)
    set_ft.sort()
    
    NT=len(set_ft)
    
    Z=np.zeros([3,NT])
    res_t=np.zeros(NT)
    #sumv=0
    for key in back_tag:
        (n,t,rt)=key   # t is time index not value
        if n==node_select:
            t=int(t)
            j=set_ft.index(rt)
            res_t[j]=rt
            Z[t,j]=back_tag[key]
                
    #        sumv+=Z[t,j]
    #print(sumv)    
    
    #chart3d = SurfaceChart3D()
    
    chart3d = BarChart3D()
    
    chart3d.style=5
    chart3d.title = "3D Bar Chart"
    chart3d.grouping='standard'
    chart3d.gapDepth=200
    chart3d.gapWidth=200
    
    for row in ws2.iter_rows():
        for cell in row:
            cell.value=None
    
    
    # rows=[]
    # for j in range(NT):
    #     rows.append((res_t[j],Z[0,j],Z[1,j],Z[2,j]))
    
    ws2['A1']=""
    i=0
    for r in ['B1','C1','D1']:
        ws2[r].value=tlist[i]
        i=i+1
    
    for row in range(NT):
        ws2.cell(row=row+2,column=1,value=res_t[row])
        for col in range(3):        
            ws2.cell(row=row+2,column=col+2,value=Z[col,row])
        
    
    data= Reference (ws2 , min_row = 1 , max_row = 3+NT, min_col = 2 , max_col = 4 )
    titles= Reference (ws2 , min_row = 2 , max_row = 3+NT, min_col = 1  )
    
    
    chart3d.add_data(data=data, titles_from_data=True)
    chart3d.set_categories(titles)
    
    
    
    ws2.add_chart(chart3d,'G2')


##----------------------------------------------------



##----------------------------------------------------event slice

NN=len(nodes)

if node_slice==0:    
    event_select=125.0  # actual time
    delt=200.0 
    #make a set fot ft
    set_ft=set(ft)
    set_ft=list(set_ft)
    set_ft.sort()
     
    Z=np.zeros([3,NN])
    
    #sumv=0
    for key in back_tag:
        (n,t,rt)=key   # t is time index not value
        if rt >= event_select and rt < event_select+delt:
            t=int(t) 
            ni=nodes.index(n)
            ni=int(ni)
            Z[t,ni]+=back_tag[key]
                
    #        sumv+=Z[t,j]
    #print(sumv)    
    
    Z=normalize(Z)
    
    #chart3d = SurfaceChart3D()
    
    chart3d = BarChart3D()
    
    chart3d.style=5
    chart3d.title = "3D Bar Chart"
    chart3d.grouping='standard'
    chart3d.gapDepth=200
    chart3d.gapWidth=200
    
    for row in ws2.iter_rows():
        for cell in row:
            cell.value=None
    
    
    ws2['A1']=""
    i=0
    for r in ['B1','C1','D1']:
        ws2[r].value=i+1
        i=i+1
    
    for row in range(NN):
        ws2.cell(row=row+2,column=1,value=nodes[row])
        for col in range(3):        
            ws2.cell(row=row+2,column=col+2,value=Z[col,row])
        
    
    data= Reference (ws2 , min_row = 1 , max_row = 3+NN, min_col = 2 , max_col = 4 )
    titles= Reference (ws2 , min_row = 2 , max_row = 3+NN, min_col = 1  )
    
    
    chart3d.add_data(data=data, titles_from_data=True)
    chart3d.set_categories(titles)
    
    
    
    ws2.add_chart(chart3d,'G2')

##----------------------------------------------------------

wb . save (file)
wb.close()

# sumv=0
# for key in back_tag:
#     (a,b,c)=key
#     if a == '2':
#         sumv+=back_tag[key]
# print(sumv)    


##--------------------------------graph draw

new_edges=[]
for edge in edges:
    e1,e2=edge
    
    e1=re.sub('\.','',e1)
    e2=re.sub('\.','',e2)
    
    if e2=='-2':
        e2=10000
    #print(edge)
    
    new_edges.append((int(e1),int(e2)))

import networkx as nk
G=nk.DiGraph(new_edges,name="Schedule")

#layout=net.kamada_kawai_layout(G)

nk.draw_networkx(G,node_size=500)#,pos=layout)

